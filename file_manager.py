import openpyxl
import pandas as pd
import datetime
import owncloud
import os
from typing import List, Tuple

class Session:
    session_name: str
    start_time: datetime.time
    def __init__(self, session_name: str, start_time_str: datetime.time) -> None:
        self.session_name = session_name
        self.start_time = start_time_str

class Sessions:
    sessions: List[Session]
    def __init__(self) -> None:
        self.sessions = []

    def add_session(self, session: Session):
        self.sessions.append(session)
    
    def get_session_names(self) -> List[str]:
        session_names: List[str] = []
        for session in self.sessions:
            session_names.append(session.session_name)
        return session_names
    
    def get_current_session(self) -> Tuple[Session, int]:
        current_datetime = datetime.datetime.now()
        current_time = current_datetime.time()
        for i in range(len(self.sessions)-1):
            if (current_time >= self.sessions[i].start_time and current_time < self.sessions[i+1].start_time):
                return self.sessions[i],i
        i = len(self.sessions)-1
        return self.sessions[i],i
    
    def get_session_by_name(self, session_name: str) -> Session:
        for i in range(len(self.sessions)):
            if self.sessions[i].session_name == session_name:
                return self.sessions[i]

class FileManager:
    workbook = openpyxl.Workbook()
    input_file_path : str
    output_file_path : str
    def __init__(self, input_file_path, output_file_path):
        self.input_file_path = input_file_path
        self.output_file_path = output_file_path
        self.oc = owncloud.Client.from_public_link('https://tuc.cloud/index.php/s/nisTB2KdNHHzgGy', folder_password = "dfg1325FDsg221")  # connect to the cloud

    def append_row_to_excel(self, value, sheet_name, session):
        current_datetime = datetime.datetime.now()
        date = current_datetime.date()
        time = current_datetime.time()
        try:
            # Load the existing Excel workbook if it exists
            workbook = openpyxl.load_workbook(self.output_file_path)
        except FileNotFoundError:
            # Create a new workbook if the file doesn't exist
            workbook = openpyxl.Workbook()

        # Check if the specified sheet already exists
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            # Create a new sheet if it doesn't exist
            sheet = workbook.create_sheet(title=sheet_name)

        # Calculate the index for the new row (after the last row)
        insert_index = sheet.max_row + 1

        # Create a DataFrame with the new row data
        data = {'Date': date, 'Time': time, 'Attendee': value, 'Session': session}
        df_row = pd.DataFrame([data])

        # Convert the DataFrame row to a list
        row_values = df_row.values.tolist()[0]

        # Insert a new row at the calculated index
        sheet.insert_rows(insert_index)

        # Populate the cells in the new row with data
        for col, value in enumerate(row_values, start=1):
            sheet.cell(row=insert_index, column=col, value=value)

        # Save the modified workbook
        workbook.save(self.output_file_path)

        # Close the workbook
        workbook.close()

    def get_sheet_names(self) -> List[str]:
        if self.input_file_path:
            wb = openpyxl.load_workbook(self.input_file_path)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        
    def appendQRData(self, value, sheetName, session):
        try:
            self.downloadRemoteFile(self.output_file_path)
        except:
            return "Error Downloading File"
        
        try:
            self.append_row_to_excel(value, sheetName, session)
        except:
            return "Error appending to File"
        
        try:
            self.upload_to_the_cloud(self.output_file_path)
            return "Saved on the cloud"
        except:
            return "Saved locally"

    def get_sessions(self, selected_sheet: str) -> List[str]:
        if selected_sheet:
            wb = openpyxl.load_workbook(self.input_file_path)
            sheet = wb[selected_sheet]
            rows = sheet.iter_rows(min_row=2, values_only=True)
            row_values = [", ".join(map(str, row)) for row in rows]
            wb.close() 
            return row_values
    
    def get_Sessions(self, selected_sheet: str) -> Sessions:
        sessions = Sessions()
        if selected_sheet:
            wb = openpyxl.load_workbook(self.input_file_path)
            sheet = wb[selected_sheet]
            rows = sheet.iter_rows(min_row=1, values_only=True)
            for row in rows:
                session_name = row[0]  # Assuming "session" is in the first column (column index 0)
                start_time = row[1]  # Assuming "start_time" is in the second column (column index 1)
                session = Session(session_name, start_time)
                sessions.add_session(session)
            wb.close()
        return sessions
        
    def downloadRemoteFile(self, saveFile: str):
        self.oc = owncloud.Client.from_public_link('https://tuc.cloud/index.php/s/nisTB2KdNHHzgGy', folder_password = "dfg1325FDsg221")  # connect to the cloud
        self.oc.get_file(saveFile, saveFile) # download remote file
        print("downloaded")
    
    def upload_to_the_cloud(self, file_name):
        self.oc.drop_file(file_name)  # upload file to cloud
        print("uploaded")
        # os.remove(file_name)     # delete local file when upload successful